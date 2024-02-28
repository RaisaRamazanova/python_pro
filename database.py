import sqlite3
from sqlite3 import Error, Connection
import random
import test_process
from data.globals import data_path
import openpyxl
import json


def create_connection(path) -> Connection | None:
    connection = None
    try:
        connection = sqlite3.connect(path)
        print("Connection to SQLite DB successful")
    except Error as e:
        print(f"The error '{e}' occurred")

    return connection


def create_tables(path):
    create_table = """
    CREATE TABLE IF NOT EXISTS questions (
        id INTEGER PRIMARY KEY,
        question TEXT,
        explanation TEXT,
        explanation_code TEXT,
        answer_1 TEXT,
        answer_2 TEXT,
        answer_3 TEXT,
        answer_4 TEXT
        );
    """
    connection = create_connection(path)
    execute_query(connection, create_table)


def execute_query(connection, query):
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        connection.commit()
        print("Query executed successfully")
    except Error as e:
        print(f"The error '{e}' occurred")


def execute_query_with_param(connection, sql, val):
    try:
        cursor = connection.cursor()
        cursor.executemany(sql, val)
        connection.commit()
        print("Query executed successfully")
        cursor.close()
    except Error as e:
        print(f"The error '{e}' occurred")
    finally:
        connection.close()
        print("MySQL connection is closed")


printed_elements = set()


def traverse_json(element, path=""):
    if isinstance(element, dict):
        for key, value in element.items():
            next_path = f"{path}.{key}" if path else key
            traverse_json(value, next_path)
    elif isinstance(element, list):
        for i, item in enumerate(element):
            next_path = f"{path}[{i}]"
            traverse_json(item, next_path)
    else:
        if element not in printed_elements:
            printed_elements.add(element)


def add_questions():
    file_path = '/Users/raisatramazanova/development/python_bot/python_pro_bot/data/stage_mapping.json'
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
        traverse_json(data)
    create_tables(data_path)

    sql = """
     INSERT INTO questions (id, question, explanation, explanation_code, answer_1, answer_2, answer_3, answer_4)
     VALUES (?, ?, ?, ?, ?, ?, ?, ?)
     """

    for element in printed_elements:
        if element == 'Node.js':
            path = f'/Users/raisatramazanova/Desktop/Data/Node.xlsx'
        else:
            path = f'/Users/raisatramazanova/Desktop/Data/{element}.xlsx'
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                row_number = 1
                for row in sheet.iter_rows(values_only=True, min_row=1):
                    if any(row) and row[0] == 0:
                        id_val = ''.join(map(str, row[1:6]))
                        val = (id_val,) + tuple(row[6:14]) + tuple([''] * (13 - len(row)))
                        sheet.cell(row=row_number, column=1).value = 1
                        data_connection = create_connection(data_path)
                        execute_query_with_param(data_connection, sql, (val,))
                    row_number += 1
            wb.save(data_path)
        except Exception as e:
            print(f"An error occurred: {e}")


async def get_question_from_data(update, context, question_id):
    questions_connection = create_connection(data_path)

    cursor = questions_connection.cursor()
    cursor.execute('SELECT id, question, explanation, explanation_code, answer_1, answer_2, answer_3, answer_4 FROM questions WHERE id = ?', (question_id,))
    data = cursor.fetchall()
    questions_connection.commit()

    last_four_elements = data[0][-4:]
    list_of_answers = [(last_four_elements[0], 1), (last_four_elements[1], 0), (last_four_elements[2], 0),
                       (last_four_elements[3], 0)]
    random.shuffle(list_of_answers)

    if str(question_id)[6] == 1:
        await test_process.four_answers_question(data[0][1], data[0][2], data[0][3], list_of_answers, question_id,
                                                 update,
                                                 context)
    else:
        await test_process.one_answers_question(data[0][1], data[0][2], question_id, update, context)